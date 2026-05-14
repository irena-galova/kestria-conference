#!/usr/bin/env python3
"""
Generate site/data/seating.json from the Singapore seating workbook.

Reads sheet "ALL" (participant roster): Member, Participant, Boats!, Practice Group,
Thursday, Friday morning, Friday afternoon.

Usage (from Agenda Website/):
    python3 parse_seating_from_xlsx.py

Requires: pip install openpyxl
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = SCRIPT_DIR / "Singapore data" / "Table setting new.xlsx"
OUT_PATH = SCRIPT_DIR / "site" / "data" / "seating.json"
SESSIONS_OUT_PATH = SCRIPT_DIR / "site" / "data" / "sessions.json"

# Maps the workbook's short session codes (Sheet1 "Session" column) to the
# human-readable label we want to show on the website. The first three entries
# below match the agenda titles for Day 2/Day 3 working sessions.
SESSION_TITLES = {
    "xbbd tf": "Cross-border BD Task Force",
    "pgs": "Practice Groups",
    "ai": "AI Ambassador Session",
    "institute": "Kestria Institute",
}

# Maps "Day" column values in Sheet1 to which slot of the My Seat card the
# session should appear under. The workbook numbers Day 1 = Wed (teambuilding,
# no tables) → Day 2 = Thursday → Day 3 = Friday.
DAY_TO_SLOT = {
    2: "thursday",      # Both XBBD TF and PGs use the participant's PG table.
    3: {                # Friday has two distinct sessions at different tables.
        "ai": "fridayAm",
        "institute": "fridayPm",
    },
}


def normalize_ws(s: str) -> str:
    return " ".join(s.split()).strip()


def parse_table_cell(val) -> int | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val != int(val):
            return int(val) if val == int(val) else None
        return int(val)
    s = normalize_ws(str(val)).lower()
    if not s or s in ("n/a", "na", "-", "—"):
        return None
    # stray newline-only cells
    if not re.search(r"\d", s):
        return None
    m = re.search(r"-?\d+", s)
    if m:
        return int(m.group(0))
    return None


def boats_slug_for_color(boats: str) -> str:
    return normalize_ws(boats).lower()


def row_to_record(row) -> dict | None:
    member = row[0]
    name = row[1]
    if name is None or (isinstance(name, str) and not normalize_ws(name)):
        return None
    if isinstance(name, str) and name.strip().lower() == "participant":
        return None

    name_clean = normalize_ws(str(name))
    # Keep known worksheet typo aligned with participant directory/search.
    if name_clean == "Elsa Camiro":
        name_clean = "Elsa Casimiro"

    boats_raw = row[2] if len(row) > 2 else None
    pg_raw = row[3] if len(row) > 3 else None
    boats = normalize_ws(str(boats_raw)) if boats_raw is not None else ""
    practice_group = normalize_ws(str(pg_raw)) if pg_raw is not None else ""

    thu = parse_table_cell(row[4]) if len(row) > 4 else None
    fri_am = parse_table_cell(row[5]) if len(row) > 5 else None
    fri_pm = parse_table_cell(row[6]) if len(row) > 6 else None

    return {
        "name": name_clean,
        "member": normalize_ws(str(member)) if member is not None else "",
        "boats": boats,
        "boatsSlug": boats_slug_for_color(boats) if boats else "",
        "practiceGroup": practice_group,
        "thursday": thu,
        "fridayAm": fri_am,
        "fridayPm": fri_pm,
    }


def clean_leader_name(val) -> str:
    """Trim a leader cell to the bare name.

    The workbook occasionally annotates leader cells with parenthetical notes
    (e.g. "Céline (if available in the meeting room)"); those are useful for
    the organisers but distracting in the participant-facing My Seat card.
    """
    if val is None:
        return ""
    s = normalize_ws(str(val))
    if "(" in s:
        s = s.split("(")[0].strip()
    return s


def parse_sessions_sheet(ws) -> dict:
    """Build the sessions/leaders structure expected by the front-end.

    Result shape:
        {
          "thursday":  [ { key, title, leaders: [t1..t6] }, ... ],
          "fridayAm":  [ ... ],
          "fridayPm":  [ ... ],
        }
    """
    by_key: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            day = int(row[0])
        except (TypeError, ValueError):
            continue
        session_raw = row[1] if len(row) > 1 else None
        if session_raw is None:
            continue
        session_key = normalize_ws(str(session_raw)).lower()
        if session_key not in SESSION_TITLES:
            continue
        table_no = parse_table_cell(row[2]) if len(row) > 2 else None
        leader = clean_leader_name(row[3] if len(row) > 3 else None)
        if table_no is None or not leader:
            continue

        slot_map = DAY_TO_SLOT.get(day)
        if slot_map is None:
            continue
        slot = slot_map[session_key] if isinstance(slot_map, dict) else slot_map

        entry_key = f"{slot}::{session_key}"
        entry = by_key.setdefault(
            entry_key,
            {
                "slot": slot,
                "key": session_key.replace(" ", "_"),
                "title": SESSION_TITLES[session_key],
                "leaders": [None] * 6,
            },
        )
        if 1 <= table_no <= 6:
            entry["leaders"][table_no - 1] = leader

    result: dict[str, list[dict]] = {"thursday": [], "fridayAm": [], "fridayPm": []}
    # Preserve a stable session order (XBBD TF before PGs on Thursday).
    slot_order = ["thursday", "fridayAm", "fridayPm"]
    session_order = ["xbbd tf", "pgs", "ai", "institute"]
    for slot in slot_order:
        for sess in session_order:
            entry = by_key.get(f"{slot}::{sess}")
            if entry:
                result[slot].append(
                    {"key": entry["key"], "title": entry["title"], "leaders": entry["leaders"]}
                )
    return result


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        print(f"Missing workbook: {xlsx}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "ALL" not in wb.sheetnames:
        print(f"No sheet 'ALL' in {xlsx}. Sheets: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb["ALL"]

    # Start from row 1 and rely on row_to_record() to skip a header row if
    # one exists. This protects us against the workbook being saved without
    # a header (so the first data row lands on row 1 and would otherwise be
    # silently dropped).
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    seating: list[dict] = []
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row[:2]):
            continue
        member = row[0]
        if isinstance(member, str) and member.strip().lower() in ("member", "member/firm", "firm"):
            continue
        rec = row_to_record(row)
        if rec:
            seating.append(rec)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as fp:
        json.dump(seating, fp, ensure_ascii=False, indent=2)
        fp.write("\n")

    print(f"Wrote {len(seating)} records to {OUT_PATH}")

    # Optional: build sessions.json from the "Sheet1" tab if present. Older
    # workbooks may not include it, so the absence is not an error.
    if "Sheet1" in wb.sheetnames:
        sessions = parse_sessions_sheet(wb["Sheet1"])
        with open(SESSIONS_OUT_PATH, "w", encoding="utf-8") as fp:
            json.dump(sessions, fp, ensure_ascii=False, indent=2)
            fp.write("\n")
        n_sess = sum(len(v) for v in sessions.values())
        print(f"Wrote {n_sess} sessions to {SESSIONS_OUT_PATH}")


if __name__ == "__main__":
    main()
