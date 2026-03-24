"""
Parse Kestria conference participant list from .xlsx into participants.json.
Only includes rows highlighted in green (fill color FF92D050).
Re-run with a new .xlsx to regenerate for a different conference.
"""

import json
import os
import re
import subprocess
import sys
import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__),
                         "2025 Budapest - MASTER FILE.xlsx")
DATA_DIR = os.path.join(os.path.dirname(__file__), "site", "data")

GREEN_FILL = "FF92D050"


def is_green(cell) -> bool:
    fg = cell.fill.fgColor
    if fg and fg.rgb and str(fg.rgb) == GREEN_FILL:
        return True
    return False


def extract_country(member: str) -> str:
    """Extract country from 'Kestria Australia' -> 'Australia'."""
    if not member:
        return ""
    m = re.match(r"Kestria\s+(.+)", member, re.IGNORECASE)
    return m.group(1).strip() if m else member.strip()


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else XLSX_PATH
    os.makedirs(DATA_DIR, exist_ok=True)

    print(f"Parsing: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Participants"]

    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers) if h}

    participants = []
    for row_idx in range(2, ws.max_row + 1):
        first_cell = ws.cell(row=row_idx, column=1)
        if not is_green(first_cell):
            continue

        def val(col_name):
            idx = col_map.get(col_name)
            if idx is None:
                return ""
            v = ws.cell(row=row_idx, column=idx + 1).value
            return str(v).strip() if v else ""

        member = val("member")
        name = val("participant")
        if not name:
            continue

        participants.append({
            "name": name,
            "member": member,
            "country": extract_country(member),
            "email": val("email"),
            "mobile": val("mobile"),
            "role": val("role") or "participant",
            "dietary": val("dietary"),
        })

    out_path = os.path.join(DATA_DIR, "participants.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(participants, f, indent=2, ensure_ascii=False)

    countries = sorted(set(p["country"] for p in participants if p["country"]))
    print(f"  -> {out_path}  ({len(participants)} participants from {len(countries)} countries)")

    # Regenerate embedded.js for offline use
    embed = os.path.join(os.path.dirname(__file__), "embed_data.py")
    if os.path.exists(embed):
        subprocess.run([sys.executable, embed], check=False)


if __name__ == "__main__":
    main()
