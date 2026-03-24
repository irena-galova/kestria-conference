"""
Parse the 'tables' sheet from 2025-Budapest-Table-Setting.xlsx into seating.json.
Each entry: { name, member, teambuilding, thuAm, thuPm, friday, fridayPG }
"""
import io
import json
import os
import subprocess
import sys
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

EXCEL = os.path.join(os.path.dirname(__file__), "2025-Budapest-Table-Setting.xlsx")
OUTPUT = os.path.join(os.path.dirname(__file__), "site", "data", "seating.json")


def safe_int(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s.lower() in ("na", "n/a", ""):
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("na", "n/a"):
        return ""
    return s


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["tables"]

    seating = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        member = safe_str(row[0]) if len(row) > 0 else ""
        name = safe_str(row[1]) if len(row) > 1 else ""
        if not name:
            continue

        entry = {
            "name": name,
            "member": member,
            "teambuilding": safe_str(row[2]) if len(row) > 2 else "",
            "thuAm": safe_int(row[3]) if len(row) > 3 else None,
            "thuPm": safe_int(row[4]) if len(row) > 4 else None,
            "friday": safe_int(row[5]) if len(row) > 5 else None,
            "fridayPG": safe_str(row[6]) if len(row) > 6 else "",
        }
        seating.append(entry)

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(seating, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(seating)} seating entries to {OUTPUT}")
    for e in seating[:3]:
        print(f"  {e['name']:30s} team={e['teambuilding']:15s} thuAm={e['thuAm']} thuPm={e['thuPm']} fri={e['friday']} pg={e['fridayPG']}")

    # Regenerate embedded.js for offline use
    embed = os.path.join(os.path.dirname(__file__), "embed_data.py")
    if os.path.exists(embed):
        subprocess.run([sys.executable, embed], check=False)


if __name__ == "__main__":
    main()
